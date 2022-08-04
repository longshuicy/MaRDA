#!/usr/bin/env python

import logging
import os
from pyclowder.extractors import Extractor
import pyclowder.files
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook


class TestPropExtractor(Extractor):
    """Count the number of characters, words and lines in a text file."""
    def __init__(self):
        Extractor.__init__(self)
        
        self.setup()

        # setup logging for the exctractor
        logging.getLogger('pyclowder').setLevel(logging.DEBUG)
        logging.getLogger('__main__').setLevel(logging.DEBUG)

    def process_message(self, connector, host, secret_key, resource, parameters):

        logger = logging.getLogger(__name__)
        inputfile = resource["local_paths"][0]
        file_id = resource['id']

        # These process messages will appear in the Clowder UI under Extractions.
        connector.message_process(resource, "Loading contents of file...")

        # 3. read data
        original_filename = resource["name"]

        if 'Results Summary' in original_filename: # trigger
            sheet_names = load_workbook(inputfile).sheetnames
            
            # bar plot
            for sheet in sheet_names:
                data = pd.read_excel(inputfile,sheet_name=sheet)
                # Get composition
                eles = data.columns[10:17]
                stoi = list(data.iloc[:,10:17].values[0])
                stoi = [round(i,2) for i in stoi]
                stoi_str = [str(i)for i in stoi]
                comp = '-'.join(['-'.join(i) for i in list(zip(*[eles,stoi_str]))])
                print(comp)

                data['Test_num'] = [i+1 for i in range(len(data))]

                fig,ax = plt.subplots(figsize=(10,5))
                # Plot Eind
                ax_x = [i-0.2 for i in data.Test_num]
                ax.bar(ax_x,data.Eind,color='k',width=0.4)
                ax.errorbar(ax_x, data.Eind, yerr = data.Eind_std, fmt ='o', c='grey',markersize=0)
                # Plot Yind
                ax2 = ax.twinx()
                ax2_x = [i+0.2 for i in data.Test_num]
                ax2.bar(ax2_x,data.Yind,color='tab:brown',width=0.4)
                ax2.errorbar(ax2_x, data.Yind, yerr = data.Yind_std, fmt ='o', c='grey',markersize=0)
                ax2.yaxis.label.set_color('tab:brown')
                ax2.spines['left'].set_color('k')
                ax2.spines['right'].set_color('tab:brown')
                ax.tick_params(axis='y', colors='k')
                ax2.tick_params(axis='y', colors='tab:brown')
                ax.set_xticks(data.Test_num)
                ax.tick_params(axis='x', which='major', labelsize=5)
                ax.set_xlabel('Test')
                ax.set_ylabel('Eind')
                ax2.set_ylabel('Yind')
                plt.xticks(rotation=45, horizontalalignment='right', fontweight='light',fontsize='x-large')
                # save figure
                figure = sheet+".jpg"
                plt.savefig(figure,dpi=300,bbox_inches='tight')
                pyclowder.files.upload_preview(connector, host, secret_key, file_id, figure)
                plt.close()
            

if __name__ == "__main__":
    extractor = TestPropExtractor()
    extractor.start()
